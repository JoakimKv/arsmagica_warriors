
# handle_warriors_class.py


import os
import pandas as pd
from openpyxl import load_workbook

from warrior_class import Warrior
from fileoperation_class import FileOperation
from random import Random


class HandleWarriors:
    
   def __init__(self, debugMode = False):

      self.debugMode = debugMode
      
      self.filename_start = r"./warriors/warriors_start.xlsx"

      if self.debugMode:

         self.filename_updated = r"./warriors/warriors_inspector_updated.xlsx"
         
      else:

         self.filename_updated = r"./warriors/warriors_updated.xlsx"

      # Create a list to store warrior objects.
      self.warriors = []

      # Create an instance of the class for fileoperations, so that the 
      # excel file may be copied.
      self.fileoperation = FileOperation()

   def getOnlyFilenameStartFromPath(self):

      path = self.filename_start
      filename = os.path.basename(path)

      return filename
   
   def getOnlyFilenameUpdatedFromPath(self):

      path = self.filename_updated
      filename = os.path.basename(path)

      return filename
   
   def loadDataFromExcelFile(self):
      
      # Load Excel file.
      self.df = pd.read_excel(self.filename_start, header = 3)

      # Clean and normalize column names.
      self.df.columns = self.df.columns.str.strip()  # Removes leading/trailing spaces.

      # print("Cleaned Columns:", self.df.columns.tolist())  # Debug output

      # Iterate through each row and create Warrior objects.
      for _, row in self.df.iterrows():

         warrior = Warrior(
            id = row["Id"],
            name = row["Name"],
            attack = row["Attack"],
            defense = row["Defense"],
            damage = row["Damage"],
            body_size = row["BodySize"],
            armour = row["Armour"],
            stance = row["Stance"],
            rating = row["Rating"],
            comments = row["Comments"]
         )
         self.warriors.append(warrior)

   def printOutWarriorList(self):
      
      # Optional: Print all warriors
      for warrior in self.warriors:        
         print(warrior)

   def updateWarriorRatings(self, addedRating):
      
      rand = Random()

      # Optional: Add rating.
      for warrior in self.warriors:        
         warrior.rating += rand.randint(-addedRating, addedRating)

   def getWarriorList(self):
      
      return self.warriors.copy()

   def setWarriorList(self, newListOfWarriors):

      bOK = False

      # Replace the current warrior list with a new one, 
      # but only if the new list is not empty.
      if newListOfWarriors and len(newListOfWarriors) > 0:

         self.warriors = []
         bOK = True

         # Make a true copy so outside changes don’t affect the class.
         for warrior in newListOfWarriors:

            self.warriors.append(warrior.makeCopyOfWarriorObject())
           
      else:
            
         print("The provided warrior list is empty in HandleWarriors class. Keeping old list.")
            
      return bOK

   def saveDataToExcelFile(self):
       
      # Load the Excel workbook
      workbook = load_workbook(self.filename_start)
      worksheet = workbook.active  # Or workbook["Sheet1"] if you know the sheet name.
    
      # Create mapping {Id: new_rating}.
      rating_updates = {warrior.id: warrior.rating for warrior in self.warriors}
    
      # Find column indexes for Id and Rating.
      header_row = 4  # Since your headers are in row 4 (Excel is 1-based).
      headers = {cell.value: cell.column for cell in worksheet[header_row]}
    
      id_col = headers["Id"]
      rating_col = headers["Rating"]
    
      # Loop through rows and update Rating.
      for row in range(header_row + 1, worksheet.max_row + 1):
         warrior_id = worksheet.cell(row = row, column = id_col).value
         if warrior_id in rating_updates:
            worksheet.cell(row = row, column = rating_col).value = rating_updates[warrior_id]
    
      # Save as updated file.
      workbook.save(self.filename_updated)
      print(f"Updated warrior ratings saved to {self.filename_updated}!")
